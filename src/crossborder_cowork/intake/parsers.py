from __future__ import annotations

import csv
import json
import mimetypes
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from ..util import stable_id


FIELD_ALIASES = {
    "product id": "product_id", "product_id": "product_id", "产品id": "product_id", "款号": "product_id", "style no": "product_id",
    "sku": "sku", "sku id": "sku", "sku_id": "sku", "货号": "sku",
    "title": "title", "product title": "title", "name": "title", "商品名称": "title", "标题": "title",
    "description": "description", "body": "description", "描述": "description",
    "category": "category", "product category": "category", "品类": "category", "类目": "category",
    "garment type": "garment_type", "type": "garment_type", "款式": "garment_type",
    "material": "materials", "materials": "materials", "材质": "materials", "面料": "materials",
    "fiber content": "fiber_content", "fiber_content": "fiber_content", "成分": "fiber_content", "纤维含量": "fiber_content",
    "care": "care_instructions", "care instructions": "care_instructions", "洗护": "care_instructions", "洗涤说明": "care_instructions",
    "country of origin": "country_of_origin", "origin": "country_of_origin", "made in": "country_of_origin", "原产国": "country_of_origin", "产地": "country_of_origin",
    "manufacturer": "manufacturer", "brand": "manufacturer", "vendor": "manufacturer", "制造商": "manufacturer", "品牌": "manufacturer",
    "color": "color", "colour": "color", "颜色": "color",
    "size": "size", "尺码": "size",
    "barcode": "barcode", "upc": "barcode", "ean": "barcode", "条码": "barcode",
    "price": "price", "价格": "price",
    "inventory": "inventory", "quantity": "inventory", "库存": "inventory",
    "images": "images", "image": "images", "image url": "images", "图片": "images",
    "tags": "tags", "标签": "tags",
    "claims": "claims", "claim": "claims", "宣传语": "claims", "卖点": "claims",
    "certifications": "certifications", "certification": "certifications", "认证": "certifications"
}


@dataclass
class ParsedRecord:
    values: dict[str, Any]
    source_document_id: str
    file_name: str
    location: str
    evidence_text: str


@dataclass
class ParsedDocument:
    id: str
    path: Path
    mime_type: str
    records: list[ParsedRecord]
    text: str = ""


def normalize_key(value: object) -> str:
    raw = str(value or "").strip()
    return FIELD_ALIASES.get(raw.casefold(), raw.casefold().replace(" ", "_").replace("-", "_"))


def normalize_row(row: dict[Any, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized = normalize_key(key)
        if isinstance(value, str):
            value = value.strip()
        if value not in (None, ""):
            result[normalized] = value
    return result


def row_text(row: dict[str, Any]) -> str:
    return " | ".join(f"{key}: {value}" for key, value in row.items())


def parse_document(path: Path) -> ParsedDocument:
    path = Path(path).resolve()
    document_id = stable_id("src", str(path), path.stat().st_size, path.stat().st_mtime_ns)
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    suffix = path.suffix.casefold()
    records: list[ParsedRecord] = []
    document_text = ""

    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            for index, raw in enumerate(csv.DictReader(stream), start=2):
                values = normalize_row(raw)
                records.append(ParsedRecord(values, document_id, path.name, f"row:{index}", row_text(values)))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                continue
            for index, values_raw in enumerate(rows, start=2):
                values = normalize_row(dict(zip(headers, values_raw)))
                if values:
                    records.append(ParsedRecord(values, document_id, path.name, f"sheet:{sheet.title};row:{index}", row_text(values)))
    elif suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        items = payload if isinstance(payload, list) else payload.get("products", payload.get("items", [payload])) if isinstance(payload, dict) else []
        for index, raw in enumerate(items, start=1):
            if isinstance(raw, dict):
                values = normalize_row(raw)
                records.append(ParsedRecord(values, document_id, path.name, f"item:{index}", row_text(values)))
    elif suffix == ".pdf":
        page_texts: list[str] = []
        for page_number, page in enumerate(PdfReader(str(path)).pages, start=1):
            text = page.extract_text() or ""
            page_texts.append(text)
            values = _parse_key_value_text(text)
            if values:
                records.append(ParsedRecord(values, document_id, path.name, f"page:{page_number}", text[:4000]))
        document_text = "\n".join(page_texts)
    elif suffix in {".md", ".markdown", ".txt"}:
        document_text = path.read_text(encoding="utf-8-sig")
        values = _parse_key_value_text(document_text)
        if values:
            records.append(ParsedRecord(values, document_id, path.name, "document", document_text[:4000]))
    elif suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
        document_text = path.name
    else:
        raise ValueError(f"Unsupported input file: {path.name}")

    return ParsedDocument(document_id, path, mime_type, records, document_text)


def _parse_key_value_text(text: str) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for line in text.splitlines():
        clean = line.strip().lstrip("#-* ")
        separator = ":" if ":" in clean else "：" if "：" in clean else ""
        if not separator:
            continue
        key, value = clean.split(separator, 1)
        if value.strip():
            values[normalize_key(key)] = value.strip()
    return values
